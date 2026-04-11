# app.py
# Main application file — contains all backend code for PestSightLog.
# Run setup.py first to create the database and default admin account.

from flask import Flask, render_template, redirect, url_for, request, flash, send_from_directory, send_file
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from flask_sqlalchemy import SQLAlchemy
from flask_wtf.csrf import CSRFProtect
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
from sqlalchemy import func
from io import BytesIO
import os
import uuid
import re
import socket
import secrets
import zipfile
import urllib.request
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def valid_email(email):
    """Check email format and verify the domain (or its parent) exists via DNS.
    Walks up subdomain levels so corporate addresses like user@purina.nestle.com
    resolve via nestle.com."""
    if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]{2,}$', email):
        return False
    parts = email.split('@')[1].lower().split('.')
    for i in range(len(parts) - 1):
        candidate = '.'.join(parts[i:])
        try:
            socket.getaddrinfo(candidate, None)
            return True
        except socket.gaierror:
            continue
    return False

EASTERN = ZoneInfo('America/New_York')

# ── App Configuration ──────────────────────────────────────────────────────────
app = Flask(__name__)

# Secret key — loaded from a local file so it is never committed to version control.
# On first run the file is created automatically with a random key.
_key_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'instance', '.secret_key')
if os.environ.get('SECRET_KEY'):
    app.config['SECRET_KEY'] = os.environ['SECRET_KEY']
else:
    os.makedirs(os.path.dirname(_key_path), exist_ok=True)
    if not os.path.exists(_key_path):
        with open(_key_path, 'w') as _f:
            _f.write(secrets.token_hex(32))
    with open(_key_path) as _f:
        app.config['SECRET_KEY'] = _f.read().strip()

# Session cookie security
app.config['SESSION_COOKIE_HTTPONLY']  = True
app.config['SESSION_COOKIE_SAMESITE']  = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)

# SQLite database stored in the same folder as this file
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///sightings.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Photo upload settings
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'uploads')
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5 MB max
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def upload_photo(file_storage):
    """Save a photo locally (fallback when Cloudinary widget is not in use).
    Returns a filename string, or None if no valid file."""
    if not file_storage or not file_storage.filename or not allowed_file(file_storage.filename):
        return None
    ext = file_storage.filename.rsplit('.', 1)[1].lower()
    filename = f'{uuid.uuid4().hex}.{ext}'
    file_storage.save(os.path.join(UPLOAD_FOLDER, filename))
    return filename

# ── Extensions Setup ───────────────────────────────────────────────────────────
db = SQLAlchemy(app)
csrf = CSRFProtect(app)

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'error'


@app.context_processor
def cloudinary_globals():
    return {
        'cloudinary_cloud_name':    os.environ.get('CLOUDINARY_CLOUD_NAME', ''),
        'cloudinary_upload_preset': os.environ.get('CLOUDINARY_UPLOAD_PRESET', ''),
    }


@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options']        = 'SAMEORIGIN'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-XSS-Protection']       = '1; mode=block'
    response.headers['Referrer-Policy']        = 'strict-origin-when-cross-origin'
    return response

@app.template_filter('localtime')
def localtime_filter(dt):
    """Convert a UTC datetime to Eastern time and format it."""
    if dt is None:
        return '—'
    return dt.replace(tzinfo=ZoneInfo('UTC')).astimezone(EASTERN).strftime('%b %d, %Y %I:%M %p %Z')


# ── Database Models ────────────────────────────────────────────────────────────

class User(UserMixin, db.Model):
    """A user account. is_admin=True gives access to all sightings."""
    id            = db.Column(db.Integer, primary_key=True)
    username      = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    is_admin      = db.Column(db.Boolean, default=False, nullable=False)
    first_name    = db.Column(db.String(80), nullable=True)
    last_name     = db.Column(db.String(80), nullable=True)
    sightings     = db.relationship('Sighting', foreign_keys='Sighting.user_id', back_populates='reporter', lazy=True)

    @property
    def full_name(self):
        if self.first_name and self.last_name:
            return f'{self.first_name} {self.last_name}'
        return self.username

    def set_password(self, password):
        """Hash and store a password."""
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        """Check a plain-text password against the stored hash."""
        return check_password_hash(self.password_hash, password)


class Sighting(db.Model):
    """A single pest sighting report submitted by a user."""
    id                = db.Column(db.Integer, primary_key=True)
    date              = db.Column(db.String(20), nullable=False)
    location          = db.Column(db.String(200), nullable=False)
    pest_type         = db.Column(db.String(100), nullable=False)
    description       = db.Column(db.Text, nullable=True)
    reported_by_name  = db.Column(db.String(100), nullable=False)
    sighting_time     = db.Column(db.String(10), nullable=True)
    reporter_email    = db.Column(db.String(200), nullable=True)
    photo_filename    = db.Column(db.String(200), nullable=True)
    status            = db.Column(db.String(20), default='open', nullable=False)
    owner_id          = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    owner_taken_at    = db.Column(db.DateTime, nullable=True)
    due_date          = db.Column(db.Date, nullable=True)

    @property
    def is_overdue(self):
        return (self.status in ('open', 'in_progress')
                and self.due_date is not None
                and self.due_date < date.today())

    @property
    def photo_url(self):
        if not self.photo_filename:
            return None
        if self.photo_filename.startswith('http'):
            return self.photo_filename
        return url_for('uploaded_file', filename=self.photo_filename)

    user_id           = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    created_at        = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    reporter          = db.relationship('User', foreign_keys=[user_id], back_populates='sightings')
    owner             = db.relationship('User', foreign_keys=[owner_id])
    capa_entries      = db.relationship('CAPAEntry', backref='sighting', lazy=True,
                                        order_by='CAPAEntry.created_at')


class CAPAEntry(db.Model):
    """A single entry in the CAPA trail for a pest sighting."""
    id             = db.Column(db.Integer, primary_key=True)
    sighting_id    = db.Column(db.Integer, db.ForeignKey('sighting.id'), nullable=False)
    user_id        = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    entry_type     = db.Column(db.String(50), nullable=False)
    description    = db.Column(db.Text, nullable=False)
    photo_filename = db.Column(db.String(200), nullable=True)
    created_at     = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    author         = db.relationship('User', foreign_keys=[user_id])

    @property
    def photo_url(self):
        if not self.photo_filename:
            return None
        if self.photo_filename.startswith('http'):
            return self.photo_filename
        return url_for('uploaded_file', filename=self.photo_filename)


# ── Flask-Login: how to load a user from their session ────────────────────────
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    today      = date.today()
    thirty_ago = datetime.utcnow() - timedelta(days=30)

    open_count        = Sighting.query.filter_by(status='open').count()
    in_progress_count = Sighting.query.filter_by(status='in_progress').count()
    completed_count   = Sighting.query.filter_by(status='completed').count()
    overdue_count     = Sighting.query.filter(
        Sighting.status.in_(['open', 'in_progress']),
        Sighting.due_date.isnot(None),
        Sighting.due_date < today
    ).count()

    recent = Sighting.query.order_by(Sighting.created_at.desc()).limit(5).all()

    top_locations = (db.session.query(Sighting.location, func.count(Sighting.id).label('cnt'))
                     .filter(Sighting.created_at >= thirty_ago)
                     .group_by(Sighting.location)
                     .order_by(func.count(Sighting.id).desc())
                     .limit(6).all())

    top_pests = (db.session.query(Sighting.pest_type, func.count(Sighting.id).label('cnt'))
                 .filter(Sighting.created_at >= thirty_ago)
                 .group_by(Sighting.pest_type)
                 .order_by(func.count(Sighting.id).desc())
                 .limit(6).all())

    return render_template('dashboard.html',
        open_count=open_count,
        in_progress_count=in_progress_count,
        completed_count=completed_count,
        overdue_count=overdue_count,
        recent=recent,
        top_locations=top_locations,
        top_pests=top_pests,
        today=today
    )


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Show the login form (GET) or process login credentials (POST)."""
    if current_user.is_authenticated:
        return redirect(url_for('sightings'))

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        user = User.query.filter(func.lower(User.username) == username.lower()).first()

        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('sightings'))
        else:
            flash('Invalid username or password.', 'error')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Log the current user out and return to login page."""
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))


@app.route('/report', methods=['GET', 'POST'])
@login_required
def report():
    """Show the report form (GET) or save a new sighting (POST)."""
    if request.method == 'POST':
        first_name    = request.form.get('first_name', '').strip()
        last_name     = request.form.get('last_name', '').strip()
        date          = request.form.get('date', '').strip()
        sighting_time = request.form.get('sighting_time', '').strip()
        location      = request.form.get('location', '').strip()
        pest_type     = request.form.get('pest_type', '').strip()
        description   = request.form.get('description', '').strip()

        reported_by_name = f'{first_name} {last_name}'.strip() or current_user.username

        if not date or not location or not pest_type or not first_name or not last_name:
            flash('First name, last name, date, location, and pest type are all required.', 'error')
        else:
            photo_filename = (request.form.get('photo_filename', '').strip()
                              or upload_photo(request.files.get('photo')))

            sighting = Sighting(
                date=date,
                sighting_time=sighting_time or None,
                location=location,
                pest_type=pest_type,
                description=description,
                reported_by_name=reported_by_name,
                photo_filename=photo_filename or None,
                user_id=current_user.id
            )
            db.session.add(sighting)
            db.session.commit()
            flash('Sighting reported successfully!', 'success')
            return redirect(url_for('sightings'))

    today = datetime.today().strftime('%Y-%m-%d')
    now_time = datetime.now().strftime('%H:%M')
    return render_template('report.html', today=today, now_time=now_time)


@app.route('/sightings')
@login_required
def sightings():
    """View all sightings with status tabs and optional search filter."""
    search          = request.args.get('search', '').strip()
    status_tab      = request.args.get('status', 'open')
    filter_location = request.args.get('filter_location', '').strip()
    filter_pest     = request.args.get('filter_pest', '').strip()

    query = Sighting.query

    if status_tab in ('open', 'in_progress', 'completed'):
        query = query.filter_by(status=status_tab)

    if filter_location:
        query = query.filter(Sighting.location == filter_location)
    elif filter_pest:
        query = query.filter(Sighting.pest_type == filter_pest)
    elif search:
        like = f'%{search}%'
        query = query.filter(
            db.or_(
                Sighting.location.ilike(like),
                Sighting.pest_type.ilike(like),
                Sighting.description.ilike(like),
                Sighting.reported_by_name.ilike(like)
            )
        )

    results = query.order_by(Sighting.created_at.desc()).all()

    counts = {
        'open':        Sighting.query.filter_by(status='open').count(),
        'in_progress': Sighting.query.filter_by(status='in_progress').count(),
        'completed':   Sighting.query.filter_by(status='completed').count(),
    }

    return render_template('sightings.html', sightings=results,
                           search=search, status_tab=status_tab, counts=counts,
                           filter_location=filter_location, filter_pest=filter_pest)


@app.route('/sightings/<int:sighting_id>')
@login_required
def sighting_detail(sighting_id):
    """Detail view for a single sighting with its CAPA trail."""
    sighting = Sighting.query.get_or_404(sighting_id)
    return render_template('sighting_detail.html', sighting=sighting, today=date.today())


@app.route('/sightings/<int:sighting_id>/take-ownership', methods=['POST'])
@login_required
def take_ownership(sighting_id):
    sighting = Sighting.query.get_or_404(sighting_id)
    if sighting.status == 'open':
        due_str = request.form.get('due_date', '').strip()
        sighting.status         = 'in_progress'
        sighting.owner_id       = current_user.id
        sighting.owner_taken_at = datetime.utcnow()
        if due_str:
            try:
                sighting.due_date = datetime.strptime(due_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        db.session.commit()
        flash('You have taken ownership of this sighting.', 'success')
    return redirect(url_for('sighting_detail', sighting_id=sighting_id))


@app.route('/sightings/<int:sighting_id>/release-ownership', methods=['POST'])
@login_required
def release_ownership(sighting_id):
    sighting = Sighting.query.get_or_404(sighting_id)
    if sighting.owner_id == current_user.id or current_user.is_admin:
        sighting.status         = 'open'
        sighting.owner_id       = None
        sighting.owner_taken_at = None
        db.session.commit()
        flash('Ownership released — sighting is back to Open.', 'success')
    return redirect(url_for('sighting_detail', sighting_id=sighting_id))


@app.route('/sightings/<int:sighting_id>/complete', methods=['POST'])
@login_required
def complete_sighting(sighting_id):
    sighting = Sighting.query.get_or_404(sighting_id)
    if sighting.owner_id == current_user.id or current_user.is_admin:
        sighting.status = 'completed'
        db.session.commit()
        flash('Sighting marked as Completed.', 'success')
    return redirect(url_for('sighting_detail', sighting_id=sighting_id))


@app.route('/sightings/<int:sighting_id>/reopen', methods=['POST'])
@login_required
def reopen_sighting(sighting_id):
    """Admin-only: reopen a completed sighting."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sighting_detail', sighting_id=sighting_id))
    sighting = Sighting.query.get_or_404(sighting_id)
    sighting.status = 'in_progress'
    db.session.commit()
    flash('Sighting reopened and set back to In Progress.', 'success')
    return redirect(url_for('sighting_detail', sighting_id=sighting_id))


@app.route('/sightings/<int:sighting_id>/capa', methods=['POST'])
@login_required
def add_capa(sighting_id):
    """Add CAPA entries to a sighting.

    form_type='response' creates up to 3 entries at once (CA + Root Cause + Preventive).
    form_type='followup' creates a single Verification or Comment entry.
    """
    sighting  = Sighting.query.get_or_404(sighting_id)
    form_type = request.form.get('form_type', 'followup')

    # Handle optional photo (widget URL takes priority, falls back to file upload)
    photo_filename = (request.form.get('photo_filename', '').strip()
                      or upload_photo(request.files.get('photo'))) or None

    if form_type == 'response':
        corrective = request.form.get('corrective', '').strip()
        root_cause = request.form.get('root_cause', '').strip()
        preventive = request.form.get('preventive', '').strip()

        if not corrective:
            flash('Corrective Action is required.', 'error')
            return redirect(url_for('sighting_detail', sighting_id=sighting_id))

        added = 0
        for entry_type, text in [
            ('Corrective Action', corrective),
            ('Root Cause',        root_cause),
            ('Preventive Action', preventive),
        ]:
            if text:
                entry = CAPAEntry(
                    sighting_id=sighting_id,
                    user_id=current_user.id,
                    entry_type=entry_type,
                    description=text,
                    photo_filename=photo_filename if entry_type == 'Corrective Action' else None
                )
                db.session.add(entry)
                added += 1

        db.session.commit()
        flash(f'CAPA response logged ({added} entr{"y" if added == 1 else "ies"} added).', 'success')

    else:  # followup
        entry_type  = request.form.get('entry_type', '').strip()
        description = request.form.get('description', '').strip()

        if not entry_type or not description:
            flash('Entry type and description are required.', 'error')
            return redirect(url_for('sighting_detail', sighting_id=sighting_id))

        entry = CAPAEntry(
            sighting_id=sighting_id,
            user_id=current_user.id,
            entry_type=entry_type,
            description=description,
            photo_filename=photo_filename
        )
        db.session.add(entry)
        db.session.commit()
        flash('Follow-up entry added.', 'success')

    return redirect(url_for('sighting_detail', sighting_id=sighting_id))


@app.route('/admin/users')
@login_required
def manage_users():
    """Admin-only page to view all user accounts."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))
    users = User.query.order_by(User.username).all()
    return render_template('users.html', users=users)


@app.route('/admin/users/add', methods=['POST'])
@login_required
def add_user():
    """Admin-only: add a new user account."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))

    username   = request.form.get('username', '').strip().lower()
    password   = request.form.get('password', '').strip()
    first_name = request.form.get('first_name', '').strip()
    last_name  = request.form.get('last_name', '').strip()
    is_admin   = request.form.get('is_admin') == 'on'

    if not username or not password or not first_name or not last_name:
        flash('All fields are required.', 'error')
    elif User.query.filter_by(username=username).first():
        flash('That username is already taken.', 'error')
    elif len(password) < 6:
        flash('Password must be at least 6 characters.', 'error')
    else:
        user = User(username=username, is_admin=is_admin,
                    first_name=first_name, last_name=last_name)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f'User "{username}" created.', 'success')

    return redirect(url_for('manage_users'))


@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
def delete_user(user_id):
    """Admin-only: delete a user account."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))

    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash('You cannot delete your own account.', 'error')
    else:
        db.session.delete(user)
        db.session.commit()
        flash(f'User "{user.username}" deleted.', 'success')

    return redirect(url_for('manage_users'))


@app.route('/submit', methods=['GET', 'POST'])
def public_report():
    """Public pest sighting form — no login required."""
    if request.method == 'POST':
        first_name    = request.form.get('first_name', '').strip()
        last_name     = request.form.get('last_name', '').strip()
        email         = request.form.get('email', '').strip()
        date_str      = request.form.get('date', '').strip()
        sighting_time = request.form.get('sighting_time', '').strip()
        location      = request.form.get('location', '').strip()
        pest_type     = request.form.get('pest_type', '').strip()
        description   = request.form.get('description', '').strip()

        if not all([first_name, last_name, email, date_str, location, pest_type]):
            return render_template('public_report.html',
                                   error='Please fill in all required fields.',
                                   today=date_str,
                                   now_time=sighting_time)

        if not valid_email(email):
            return render_template('public_report.html',
                                   error='Please enter a valid email address (e.g. name@example.com).',
                                   today=date_str,
                                   now_time=sighting_time)

        photo_filename = (request.form.get('photo_filename', '').strip()
                          or upload_photo(request.files.get('photo'))) or None

        # Use the hidden system user for public submissions
        system_user = User.query.filter_by(username='_public_').first()

        sighting = Sighting(
            date=date_str,
            sighting_time=sighting_time or None,
            location=location,
            pest_type=pest_type,
            description=description,
            reported_by_name=f'{first_name} {last_name}',
            reporter_email=email,
            photo_filename=photo_filename,
            user_id=system_user.id
        )
        db.session.add(sighting)
        db.session.commit()
        return render_template('public_report.html',
                               success=True,
                               today=datetime.today().strftime('%Y-%m-%d'),
                               now_time=datetime.now().strftime('%H:%M'))

    today    = datetime.today().strftime('%Y-%m-%d')
    now_time = datetime.now().strftime('%H:%M')
    return render_template('public_report.html', today=today, now_time=now_time)


@app.route('/sightings/<int:sighting_id>/delete', methods=['POST'])
@login_required
def delete_sighting(sighting_id):
    """Admin-only: permanently delete a sighting and its CAPA entries."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))
    sighting = Sighting.query.get_or_404(sighting_id)
    CAPAEntry.query.filter_by(sighting_id=sighting_id).delete()
    db.session.delete(sighting)
    db.session.commit()
    flash(f'Sighting #{sighting_id} has been deleted.', 'success')
    return redirect(url_for('sightings', status='open'))


@app.route('/uploads/<filename>')
def uploaded_file(filename):
    """Serve an uploaded photo."""
    safe_name = os.path.basename(filename)
    return send_from_directory(UPLOAD_FOLDER, safe_name)


@app.route('/admin/users/<int:user_id>/edit-name', methods=['POST'])
@login_required
def edit_user_name(user_id):
    """Admin-only: update a user's first and last name."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))
    user = User.query.get_or_404(user_id)
    user.first_name = request.form.get('first_name', '').strip()
    user.last_name  = request.form.get('last_name', '').strip()
    db.session.commit()
    flash(f'Name updated for "{user.username}".', 'success')
    return redirect(url_for('manage_users'))


@app.route('/admin/users/<int:user_id>/set-password', methods=['POST'])
@login_required
def set_user_password(user_id):
    """Admin-only: reset another user's password."""
    if not current_user.is_admin:
        flash('Admin access required.', 'error')
        return redirect(url_for('sightings'))
    user = User.query.get_or_404(user_id)
    new_password = request.form.get('new_password', '').strip()
    if len(new_password) < 6:
        flash('Password must be at least 6 characters.', 'error')
    else:
        user.set_password(new_password)
        db.session.commit()
        flash(f'Password for "{user.username}" has been updated.', 'success')
    return redirect(url_for('manage_users'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """Allow any logged-in user to change their own password."""
    if request.method == 'POST':
        current_pw = request.form.get('current_password', '')
        new_pw     = request.form.get('new_password', '').strip()
        confirm_pw = request.form.get('confirm_password', '').strip()

        if not current_user.check_password(current_pw):
            flash('Current password is incorrect.', 'error')
        elif len(new_pw) < 6:
            flash('New password must be at least 6 characters.', 'error')
        elif new_pw != confirm_pw:
            flash('New passwords do not match.', 'error')
        else:
            current_user.set_password(new_pw)
            db.session.commit()
            flash('Password changed successfully.', 'success')
            return redirect(url_for('sightings'))

    return render_template('change_password.html')


# ── Export Routes ──────────────────────────────────────────────────────────────

@app.route('/export')
@login_required
def export():
    return render_template('export.html', today=date.today())


def _full_photo_url(photo_filename):
    """Return a fully-qualified URL for a photo — works for both Cloudinary URLs
    and old locally-stored filenames."""
    if not photo_filename:
        return None
    if photo_filename.startswith('http'):
        return photo_filename
    return url_for('uploaded_file', filename=photo_filename, _external=True)


def _build_export_query():
    """Build a Sighting query from request args (from_date, to_date, status)."""
    from_date = request.args.get('from_date', '').strip()
    to_date   = request.args.get('to_date', '').strip()
    statuses  = request.args.getlist('status')

    query = Sighting.query
    if from_date:
        query = query.filter(Sighting.date >= from_date)
    if to_date:
        query = query.filter(Sighting.date <= to_date)
    if statuses:
        query = query.filter(Sighting.status.in_(statuses))
    return query.order_by(Sighting.date.desc()).all()


@app.route('/export/excel')
@login_required
def export_excel():
    sightings = _build_export_query()

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='052818')
    header_align = Alignment(horizontal='center')

    # ── Sheet 1: Sightings ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Sightings'
    cols1 = ['ID', 'Date', 'Time', 'Pest Type', 'Location', 'Reported By',
             'Email', 'Status', 'Owner', 'Due Date', 'Overdue', 'Description', 'Photo']
    for c, h in enumerate(cols1, 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for r, s in enumerate(sightings, 2):
        ws1.cell(row=r, column=1,  value=s.id)
        ws1.cell(row=r, column=2,  value=s.date)
        ws1.cell(row=r, column=3,  value=s.sighting_time or '')
        ws1.cell(row=r, column=4,  value=s.pest_type)
        ws1.cell(row=r, column=5,  value=s.location)
        ws1.cell(row=r, column=6,  value=s.reported_by_name)
        ws1.cell(row=r, column=7,  value=s.reporter_email or '')
        status_label = {'open': 'Needs Action', 'in_progress': 'In Progress',
                        'completed': 'Completed'}.get(s.status, s.status)
        ws1.cell(row=r, column=8,  value=status_label)
        ws1.cell(row=r, column=9,  value=s.owner.full_name if s.owner else '')
        ws1.cell(row=r, column=10, value=str(s.due_date) if s.due_date else '')
        ws1.cell(row=r, column=11, value='Yes' if s.is_overdue else 'No')
        ws1.cell(row=r, column=12, value=s.description or '')
        _url = _full_photo_url(s.photo_filename)
        if _url:
            cell = ws1.cell(row=r, column=13, value=_url)
            cell.hyperlink = _url
            cell.font = Font(color='0563C1', underline='single')

    # ── Sheet 2: CAPA Trail ───────────────────────────────────────────────────
    ws2 = wb.create_sheet('CAPA Trail')
    cols2 = ['Sighting ID', 'Pest Type', 'Location', 'Entry ID', 'Date', 'Time',
             'Entry Type', 'Author', 'Description', 'Photo']
    for c, h in enumerate(cols2, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Alternating fill colours for each sighting group
    fills = [
        PatternFill(fill_type='solid', fgColor='F0F7F4'),  # light green
        PatternFill(fill_type='solid', fgColor='FFFFFF'),  # white
    ]

    row = 2
    for group_idx, s in enumerate(sightings):
        if not s.capa_entries:
            continue
        row_fill = fills[group_idx % 2]
        for entry in s.capa_entries:
            local_dt = entry.created_at.replace(tzinfo=ZoneInfo('UTC')).astimezone(EASTERN)
            values = [
                s.id,
                s.pest_type,
                s.location,
                entry.id,
                local_dt.strftime('%Y-%m-%d'),
                local_dt.strftime('%I:%M %p'),
                entry.entry_type,
                entry.author.full_name,
                entry.description,
            ]
            for c, val in enumerate(values, 1):
                cell = ws2.cell(row=row, column=c, value=val)
                cell.fill = row_fill
            _url = _full_photo_url(entry.photo_filename)
            if _url:
                cell = ws2.cell(row=row, column=10, value=_url)
                cell.hyperlink = _url
                cell.font = Font(color='0563C1', underline='single')
                cell.fill = row_fill
            row += 1
        # Blank divider row between sighting groups
        row += 1

    # Auto-size columns
    for ws in [ws1, ws2]:
        for col in ws.columns:
            width = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(width + 4, 60)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'pestsightlog_{date.today()}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


@app.route('/export/photos')
@login_required
def export_photos():
    sightings = _build_export_query()

    output = BytesIO()
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zf:
        for s in sightings:
            def _add_photo(photo_ref, zip_path):
                if not photo_ref:
                    return
                if photo_ref.startswith('http'):
                    try:
                        with urllib.request.urlopen(photo_ref, timeout=10) as resp:
                            ext = photo_ref.split('.')[-1].split('?')[0][:5]
                            zf.writestr(f'{zip_path}.{ext}', resp.read())
                    except Exception:
                        pass
                else:
                    local = os.path.join(UPLOAD_FOLDER, photo_ref)
                    if os.path.exists(local):
                        ext = photo_ref.rsplit('.', 1)[-1]
                        zf.write(local, f'{zip_path}.{ext}')

            _add_photo(s.photo_filename, f'sighting_{s.id}/sighting_photo')
            for entry in s.capa_entries:
                _add_photo(entry.photo_filename,
                           f'sighting_{s.id}/capa_entry_{entry.id}_photo')

    output.seek(0)
    filename = f'pestsightlog_photos_{date.today()}.zip'
    return send_file(output, mimetype='application/zip',
                     as_attachment=True, download_name=filename)


# ── Entry Point ────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    app.run(debug=False)
